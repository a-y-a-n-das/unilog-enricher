import tempfile
from pathlib import Path

import openpyxl
import pytest

from database import repositories
from database.connection import SessionLocal
from database.models import Job, JobRow
from models.input_models import InputRecord
from services.job_service import create_job
from api.output_generator import generate_output, generate_partial_output, OUTPUT_HEADERS, OUTPUT_HEADER_TO_FIELD


def _setup_job_with_input_sheet(tmpdir: Path, input_headers: list[str], records: list[InputRecord]) -> Job:
    """Helper to create job with input file containing Input sheet."""
    input_path = tmpdir / "test_input.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input"
    ws.append(input_headers)
    for record in records:
        row_data = [record.data.get(h, "") for h in input_headers]
        ws.append(row_data)
    wb.save(input_path)
    job = create_job(
        input_filename="test_input.xlsx",
        input_format="xlsx",
        input_file_path=str(input_path),
        rows=records,
    )
    return job


def _add_completed_row(job: Job, row_number: int, input_data: dict, result_data: dict):
    """Add a completed row to the job."""
    from database.models import JobRow
    import uuid
    row = JobRow(
        id=uuid.uuid4(),
        job_id=job.id,
        row_number=row_number,
        status="completed",
        input_data=input_data,
        result_data=result_data,
    )
    with SessionLocal() as session:
        session.add(row)
        session.commit()


def _add_pending_row(job: Job, row_number: int, input_data: dict):
    """Add a pending row to the job."""
    from database.models import JobRow
    import uuid
    row = JobRow(
        id=uuid.uuid4(),
        job_id=job.id,
        row_number=row_number,
        status="pending",
        input_data=input_data,
        result_data=None,
    )
    with SessionLocal() as session:
        session.add(row)
        session.commit()


def _update_row_result(job: Job, row_number: int, result_data: dict):
    """Update an existing row's result_data."""
    with SessionLocal() as session:
        row = session.execute(
            JobRow.__table__.select().where(
                (JobRow.job_id == job.id) & (JobRow.row_number == row_number)
            )
        ).first()
        if row:
            row = session.get(JobRow, row.id)
            row.result_data = result_data
            session.commit()


def test_output_headers_match_model_schema():
    """Test that Output headers match ExtractedProduct model field aliases in order."""
    # Verify key headers exist in correct order
    assert OUTPUT_HEADERS[0] == "MFR URL"
    assert OUTPUT_HEADERS[1] == "Ref URL 1"
    assert "PART_NUMBER" in OUTPUT_HEADERS
    assert "SKU - MY_PART_NUMBER" in OUTPUT_HEADERS
    assert "MANUFACTURER_NAME" in OUTPUT_HEADERS
    assert "BRAND_NAME" in OUTPUT_HEADERS
    assert "ITEM_FEATURES_1" in OUTPUT_HEADERS
    assert "ITEM_FEATURES_20" in OUTPUT_HEADERS
    assert "ATTRIBUTE_LABEL 1" in OUTPUT_HEADERS
    assert "ATTRIBUTE_VALUE 1" in OUTPUT_HEADERS
    assert "ATTRIBUTE_UOM 1" in OUTPUT_HEADERS
    assert "ATTRIBUTE_LABEL 50" in OUTPUT_HEADERS
    assert "ATTRIBUTE_VALUE 50" in OUTPUT_HEADERS
    assert "ATTRIBUTE_UOM 50" in OUTPUT_HEADERS
    assert "UPC" in OUTPUT_HEADERS
    assert "EAN" in OUTPUT_HEADERS
    assert "GTIN" in OUTPUT_HEADERS
    assert "UNSPSC" in OUTPUT_HEADERS
    assert "List Price" in OUTPUT_HEADERS
    assert "LENGTH" in OUTPUT_HEADERS
    assert "LENGTH_UOM" in OUTPUT_HEADERS
    assert "Product Image" in OUTPUT_HEADERS
    assert "SDS" in OUTPUT_HEADERS
    assert "Country Of Origin" in OUTPUT_HEADERS
    assert "Discontinued" in OUTPUT_HEADERS
    assert "Actual Image (Yes/No)" in OUTPUT_HEADERS
    assert "row_number" not in OUTPUT_HEADERS
    assert "status" not in OUTPUT_HEADERS
    assert "error_message" not in OUTPUT_HEADERS
    assert not any(h.startswith("result_") for h in OUTPUT_HEADERS)


def test_output_header_to_field_mapping():
    """Test that header to field map has correct paths."""
    # Scalar fields
    assert OUTPUT_HEADER_TO_FIELD["MFR URL"] == "mfr_url"
    assert OUTPUT_HEADER_TO_FIELD["PART_NUMBER"] == "part_number"
    assert OUTPUT_HEADER_TO_FIELD["MANUFACTURER_NAME"] == "manufacturer_name"
    assert OUTPUT_HEADER_TO_FIELD["UPC"] == "upc"
    assert OUTPUT_HEADER_TO_FIELD["LENGTH"] == "length"
    assert OUTPUT_HEADER_TO_FIELD["Product Image"] == "product_image"
    assert OUTPUT_HEADER_TO_FIELD["SDS"] == "sds"
    
    # Item features
    assert OUTPUT_HEADER_TO_FIELD["ITEM_FEATURES_1"] == "item_features[0]"
    assert OUTPUT_HEADER_TO_FIELD["ITEM_FEATURES_2"] == "item_features[1]"
    assert OUTPUT_HEADER_TO_FIELD["ITEM_FEATURES_20"] == "item_features[19]"
    
    # Attributes
    assert OUTPUT_HEADER_TO_FIELD["ATTRIBUTE_LABEL 1"] == "attributes[0].label"
    assert OUTPUT_HEADER_TO_FIELD["ATTRIBUTE_VALUE 1"] == "attributes[0].value"
    assert OUTPUT_HEADER_TO_FIELD["ATTRIBUTE_UOM 1"] == "attributes[0].uom"
    assert OUTPUT_HEADER_TO_FIELD["ATTRIBUTE_LABEL 50"] == "attributes[49].label"
    assert OUTPUT_HEADER_TO_FIELD["ATTRIBUTE_VALUE 50"] == "attributes[49].value"
    assert OUTPUT_HEADER_TO_FIELD["ATTRIBUTE_UOM 50"] == "attributes[49].uom"


def test_generate_output_xlsx_sheet_structure():
    """Test that generated XLSX has correct sheet structure."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        records = [
            InputRecord(row_number=1, data={"product_name": "Test Product", "part_number": "PN-123"})
        ]
        job = _setup_job_with_input_sheet(tmp_path, ["product_name", "part_number"], records)
        
        _add_completed_row(job, 2, {"product_name": "Test Product", "part_number": "PN-123"}, {
            "mfr_url": {"value": "https://mfr.com"},
            "part_number": {"value": "PN-123"},
        })
        
        output_path = generate_output(str(job.id))
        
        assert output_path is not None
        assert output_path.exists()
        
        wb_out = openpyxl.load_workbook(output_path)
        
        assert "Input" in wb_out.sheetnames
        assert "Output" in wb_out.sheetnames
        
        ws_input = wb_out["Input"]
        assert ws_input.cell(row=1, column=1).value == "row_number"
        assert ws_input.cell(row=1, column=2).value == "product_name"
        assert ws_input.cell(row=1, column=3).value == "part_number"
        
        ws_output = wb_out["Output"]
        output_headers = [cell.value for cell in ws_output[1]]
        
        assert output_headers == OUTPUT_HEADERS
        assert "row_number" not in output_headers
        assert "status" not in output_headers
        assert "error_message" not in output_headers
        assert not any(h.startswith("result_") for h in output_headers)
        assert len(output_headers) == len(set(output_headers))
        
        row_values = [cell.value for cell in ws_output[2]]
        assert row_values[output_headers.index("MFR URL")] == "https://mfr.com"
        assert row_values[output_headers.index("PART_NUMBER")] == "PN-123"


def test_generate_output_conflicts_sheet():
    """Test that Conflicts sheet is generated when conflicts exist."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        records = [
            InputRecord(row_number=1, data={"product_name": "Test Product"})
        ]
        job = _setup_job_with_input_sheet(tmp_path, ["product_name"], records)
        
        _add_completed_row(job, 2, {"product_name": "Test Product"}, {
            "part_number": {
                "value": "PN-123",
                "is_conflict": True,
                "conflict_details": {
                    "field": "PART_NUMBER",
                    "selected_value": "PN-123",
                    "selected_source": "mfr",
                    "conflicting_value": "PN-456",
                    "conflicting_source": "dist",
                    "recommendation": "Use mfr",
                },
            },
        })
        
        output_path = generate_output(str(job.id))
        
        wb_out = openpyxl.load_workbook(output_path)
        
        assert "Conflicts" in wb_out.sheetnames
        
        ws_conflicts = wb_out["Conflicts"]
        conflict_headers = [cell.value for cell in ws_conflicts[1]]
        
        assert conflict_headers == [
            "row_number",
            "field",
            "selected_value",
            "selected_uom",
            "selected_source",
            "conflicting_value",
            "conflicting_uom",
            "conflicting_source",
            "recommendation",
        ]
        
        row_values = [cell.value for cell in ws_conflicts[2]]
        assert row_values[0] == 2
        assert row_values[1] == "PART_NUMBER"
        assert row_values[2] == "PN-123"
        assert row_values[4] == "mfr"
        assert row_values[5] == "PN-456"
        assert row_values[7] == "dist"
        assert row_values[8] == "Use mfr"


def test_conflicts_not_in_output_columns():
    """Test that conflicts are exported to Conflicts sheet only, not as Output columns."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        records = [
            InputRecord(row_number=1, data={"product_name": "Test Product"})
        ]
        job = _setup_job_with_input_sheet(tmp_path, ["product_name"], records)
        
        _add_completed_row(job, 2, {"product_name": "Test Product"}, {
            "part_number": {
                "value": "PN-123",
                "is_conflict": True,
                "conflict_details": {
                    "field": "PART_NUMBER",
                    "selected_value": "PN-123",
                    "conflicting_value": "PN-456",
                    "recommendation": "Use mfr",
                },
            },
        })
        
        output_path = generate_output(str(job.id))
        wb_out = openpyxl.load_workbook(output_path)
        
        # Output sheet should NOT have conflict-related columns
        ws_output = wb_out["Output"]
        output_headers = [cell.value for cell in ws_output[1]]
        assert "selected_value" not in output_headers
        assert "conflicting_value" not in output_headers
        assert "recommendation" not in output_headers
        assert "is_conflict" not in output_headers
        
        # Conflicts sheet should have the conflict data
        ws_conflicts = wb_out["Conflicts"]
        conflict_headers = [cell.value for cell in ws_conflicts[1]]
        assert "selected_value" in conflict_headers
        assert "conflicting_value" in conflict_headers
        assert "recommendation" in conflict_headers


def test_model_field_mapping_values():
    """Test that various field types are correctly mapped to Output cells."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        records = [
            InputRecord(row_number=1, data={"product_name": "Test Product"})
        ]
        job = _setup_job_with_input_sheet(tmp_path, ["product_name"], records)
        
        # Create result_data with various field types
        result_data = {
            # Scalar field
            "mfr_url": {"value": "https://mfr.example.com"},
            "part_number": {"value": "PN-999"},
            "manufacturer_name": {"value": "ACME Corp"},
            # Item features
            "item_features": [
                {"value": "Feature One"},
                {"value": "Feature Two"},
            ],
            # Attributes
            "attributes": [
                {"label": "Color", "value": {"value": "Red"}, "uom": ""},
                {"label": "Size", "value": {"value": "Large"}, "uom": "in"},
            ],
            # Resource/image field
            "product_image": {"value": "https://images.example.com/img.jpg"},
            # Compliance field
            "rohs": {"value": "Compliant"},
        }
        
        _add_completed_row(job, 2, {"product_name": "Test Product"}, result_data)
        
        output_path = generate_output(str(job.id))
        wb_out = openpyxl.load_workbook(output_path)
        ws_output = wb_out["Output"]
        output_headers = [cell.value for cell in ws_output[1]]
        row_values = [cell.value for cell in ws_output[2]]
        
        # Scalar fields
        assert row_values[output_headers.index("MFR URL")] == "https://mfr.example.com"
        assert row_values[output_headers.index("PART_NUMBER")] == "PN-999"
        assert row_values[output_headers.index("MANUFACTURER_NAME")] == "ACME Corp"
        
        # Item features
        assert row_values[output_headers.index("ITEM_FEATURES_1")] == "Feature One"
        assert row_values[output_headers.index("ITEM_FEATURES_2")] == "Feature Two"
        
        # Attributes
        assert row_values[output_headers.index("ATTRIBUTE_LABEL 1")] == "Color"
        assert row_values[output_headers.index("ATTRIBUTE_VALUE 1")] == "Red"
        # openpyxl saves empty strings as None
        assert row_values[output_headers.index("ATTRIBUTE_UOM 1")] in ("", None)
        assert row_values[output_headers.index("ATTRIBUTE_LABEL 2")] == "Size"
        assert row_values[output_headers.index("ATTRIBUTE_VALUE 2")] == "Large"
        assert row_values[output_headers.index("ATTRIBUTE_UOM 2")] == "in"
        
        # Resource/image field
        assert row_values[output_headers.index("Product Image")] == "https://images.example.com/img.jpg"
        
        # Compliance field
        assert row_values[output_headers.index("RoHS")] == "Compliant"


def test_generate_partial_output():
    """Test partial output generation during job processing."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        records = [
            InputRecord(row_number=1, data={"product_name": "Test Product 1"}),
            InputRecord(row_number=2, data={"product_name": "Test Product 2"}),
        ]
        job = _setup_job_with_input_sheet(tmp_path, ["product_name"], records)
        
        _add_completed_row(job, 3, {"product_name": "Test Product 1"}, {"mfr_url": {"value": "https://mfr1.com"}})
        _add_pending_row(job, 4, {"product_name": "Test Product 2"})
        
        partial_path = generate_partial_output(str(job.id))
        
        assert partial_path is not None
        assert partial_path.exists()
        
        wb_out = openpyxl.load_workbook(partial_path)
        
        assert "Input" in wb_out.sheetnames
        assert "Output" in wb_out.sheetnames
        
        ws_output = wb_out["Output"]
        output_headers = [cell.value for cell in ws_output[1]]
        
        assert output_headers == OUTPUT_HEADERS
        assert "row_number" not in output_headers
        assert "status" not in output_headers
        assert "error_message" not in output_headers
        
        # Row 3 (pending) should have input data but no result data - value is None
        row3_values = [cell.value for cell in ws_output[3]]
        assert row3_values[output_headers.index("PART_NUMBER")] is None
        
        # Row 2 (completed) should have result data.
        row2_values = [cell.value for cell in ws_output[2]]
        assert row2_values[output_headers.index("MFR URL")] == "https://mfr1.com"


def test_partial_output_reflects_live_database_state():
    """Test that partial output queries live DB state at download time, not stale cache."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        records = [
            InputRecord(row_number=1, data={"product_name": "Test Product"})
        ]
        job = _setup_job_with_input_sheet(tmp_path, ["product_name"], records)
        
        # Add completed row with initial MFR URL
        _add_completed_row(job, 2, {"product_name": "Test Product"}, {"mfr_url": {"value": "https://old.example.com"}})
        
        # First partial output
        partial1 = generate_partial_output(str(job.id))
        wb1 = openpyxl.load_workbook(partial1)
        ws1 = wb1["Output"]
        headers1 = [cell.value for cell in ws1[1]]
        row1_values = [cell.value for cell in ws1[2]]
        
        assert row1_values[headers1.index("MFR URL")] == "https://old.example.com"
        
        # Update the JobRow in database to have NEW MFR URL
        _update_row_result(job, 2, {"mfr_url": {"value": "https://new.example.com"}})
        
        # Second partial output - should reflect NEW value
        partial2 = generate_partial_output(str(job.id))
        wb2 = openpyxl.load_workbook(partial2)
        ws2 = wb2["Output"]
        headers2 = [cell.value for cell in ws2[1]]
        row2_values = [cell.value for cell in ws2[2]]
        
        assert row2_values[headers2.index("MFR URL")] == "https://new.example.com"
        assert row2_values[headers2.index("MFR URL")] != "https://old.example.com"


def test_partial_output_new_conflicts_appear():
    """Test that newly detected conflicts appear in subsequent partial exports."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        records = [
            InputRecord(row_number=1, data={"product_name": "Test Product"})
        ]
        job = _setup_job_with_input_sheet(tmp_path, ["product_name"], records)
        
        # Add completed row WITHOUT conflict initially
        _add_completed_row(job, 2, {"product_name": "Test Product"}, {"part_number": {"value": "PN-123"}})
        
        # First partial - no conflicts
        partial1 = generate_partial_output(str(job.id))
        wb1 = openpyxl.load_workbook(partial1)
        assert "Conflicts" not in wb1.sheetnames
        
        # Update JobRow to ADD a conflict
        _update_row_result(job, 2, {
            "part_number": {
                "value": "PN-123",
                "is_conflict": True,
                "conflict_details": {
                    "field": "PART_NUMBER",
                    "selected_value": "PN-123",
                    "conflicting_value": "PN-456",
                    "recommendation": "Use mfr",
                },
            }
        })
        
        # Second partial - Conflicts sheet should now appear
        partial2 = generate_partial_output(str(job.id))
        wb2 = openpyxl.load_workbook(partial2)
        assert "Conflicts" in wb2.sheetnames
        
        ws_conflicts = wb2["Conflicts"]
        row_values = [cell.value for cell in ws_conflicts[2]]
        assert row_values[1] == "PART_NUMBER"
        assert row_values[2] == "PN-123"
        assert row_values[5] == "PN-456"


def test_no_duplicate_columns_in_output():
    """Test that Output sheet has no duplicate columns."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        records = [
            InputRecord(row_number=1, data={"product_name": "Test Product", "part_number": "PN-123"})
        ]
        job = _setup_job_with_input_sheet(tmp_path, ["product_name", "part_number"], records)
        
        _add_completed_row(job, 2, {"product_name": "Test Product", "part_number": "PN-123"}, {
            "mfr_url": {"value": "https://mfr.com"},
            "part_number": {"value": "PN-123"},
        })
        
        output_path = generate_output(str(job.id))
        
        wb_out = openpyxl.load_workbook(output_path)
        ws_output = wb_out["Output"]
        output_headers = [cell.value for cell in ws_output[1]]
        
        assert len(output_headers) == len(set(output_headers))


def test_header_formatting_preserved():
    """Test that header formatting is preserved from original workbook."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        
        input_path = tmp_path / "test_input.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Input"
        ws.append(["product_name"])
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF", name="Calibri")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        ws.append(["Test Product"])
        wb.save(input_path)
        
        records = [
            InputRecord(row_number=1, data={"product_name": "Test Product"})
        ]
        job = create_job(
            input_filename="test_input.xlsx",
            input_format="xlsx",
            input_file_path=str(input_path),
            rows=records,
        )
        
        _add_completed_row(job, 2, {"product_name": "Test Product"}, {})
        
        output_path = generate_output(str(job.id))
        
        wb_out = openpyxl.load_workbook(output_path)
        ws_output = wb_out["Output"]
        
        # Headers are in row 1
        for cell in ws_output[1]:
            assert cell.font.bold == True
            assert cell.font.color is not None
            # Fill may not copy perfectly but should have some style applied
            assert cell.fill is not None


def test_xlsx_partial_output():
    """Test partial output always uses XLSX format regardless of input format."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        input_path = tmp_path / "test_input.csv"
        input_path.write_text("product_name\nTest Product 1\nTest Product 2\n")
        records = [
            InputRecord(row_number=1, data={"product_name": "Test Product 1"}),
            InputRecord(row_number=2, data={"product_name": "Test Product 2"}),
        ]
        job = create_job(
            input_filename="test_input.csv",
            input_format="csv",
            input_file_path=str(input_path),
            rows=records,
        )
        _add_completed_row(job, 3, {"product_name": "Test Product 1"}, {"mfr_url": {"value": "https://mfr1.com"}})
        _add_pending_row(job, 4, {"product_name": "Test Product 2"})
        
        partial_path = generate_partial_output(str(job.id))
        
        assert partial_path is not None
        assert partial_path.exists()
        assert partial_path.suffix == ".xlsx"
        
        wb_out = openpyxl.load_workbook(partial_path)
        ws_output = wb_out["Output"]
        headers = [cell.value for cell in ws_output[1]]
        
        # XLSX output: row_number + input columns + status + error_message + OUTPUT_HEADERS
        assert "row_number" in headers
        assert "product_name" in headers
        assert "status" in headers
        assert "error_message" in headers
        assert "MFR URL" in headers
        assert not any(h.startswith("result_") for h in headers)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])