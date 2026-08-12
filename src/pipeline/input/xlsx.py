from pathlib import Path
from typing import Any

import openpyxl

from models.input import InputRecord


REQUIRED_SHEET = "Input"


def load_input_xlsx(path: str | Path) -> list[InputRecord]:
    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    workbook = openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    if REQUIRED_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"Workbook must contain a '{REQUIRED_SHEET}' sheet"
        )

    worksheet = workbook[REQUIRED_SHEET]

    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        raise ValueError("Input sheet is empty")

    headers = [
        str(header).strip() if header is not None else ""
        for header in headers
    ]

    if any(not header for header in headers):
        raise ValueError("Input sheet contains an empty column header")

    records: list[InputRecord] = []

    for row_number, row in enumerate(rows, start=2):
        if not any(value is not None for value in row):
            continue

        data: dict[str, Any] = dict(zip(headers, row))

        records.append(
            InputRecord(
                row_number=row_number,
                data=data,
            )
        )

    return records
