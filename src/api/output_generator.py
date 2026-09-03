from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from database import repositories
from database.connection import SessionLocal
from database.models import Job, JobRow
from models.extraction_models import ExtractedProduct

from api.storage import get_output_file_path


def _get_model_alias_to_field_map() -> dict[str, str]:
    """Map model alias -> ExtractedProduct field name."""
    mapping: dict[str, str] = {}

    for field_name, field_info in ExtractedProduct.model_fields.items():
        if field_name == "row_number":
            continue

        alias = field_info.alias

        # item_features and attributes don't have aliases but expand to multiple headers
        if field_name == "item_features":
            for i in range(1, 21):
                mapping[f"ITEM_FEATURES_{i}"] = f"item_features[{i - 1}]"

        elif field_name == "attributes":
            for i in range(1, 51):
                mapping[f"ATTRIBUTE_LABEL {i}"] = f"attributes[{i - 1}].label"
                mapping[f"ATTRIBUTE_VALUE {i}"] = f"attributes[{i - 1}].value"
                mapping[f"ATTRIBUTE_UOM {i}"] = f"attributes[{i - 1}].uom"

        elif alias:
            mapping[alias] = field_name

    return mapping


MODEL_ALIAS_TO_FIELD = _get_model_alias_to_field_map()


def _get_output_headers() -> list[str]:
    """Get the fixed Output schema headers from ExtractedProduct model in definition order."""
    headers: list[str] = []

    for field_name, field_info in ExtractedProduct.model_fields.items():
        if field_name == "row_number":
            continue

        alias = field_info.alias

        if field_name == "item_features":
            for i in range(1, 21):
                headers.append(f"ITEM_FEATURES_{i}")

        elif field_name == "attributes":
            for i in range(1, 51):
                headers.append(f"ATTRIBUTE_LABEL {i}")
                headers.append(f"ATTRIBUTE_VALUE {i}")
                headers.append(f"ATTRIBUTE_UOM {i}")

        elif alias:
            headers.append(alias)

    return headers


def _get_output_header_to_field_map() -> dict[str, str]:
    """Map Output header name to ExtractedProduct field name."""
    mapping: dict[str, str] = {}

    for field_name, field_info in ExtractedProduct.model_fields.items():
        if field_name == "row_number":
            continue

        alias = field_info.alias

        # item_features and attributes don't have aliases but expand to multiple headers
        if field_name == "item_features":
            for i in range(1, 21):
                mapping[f"ITEM_FEATURES_{i}"] = f"item_features[{i - 1}]"

        elif field_name == "attributes":
            for i in range(1, 51):
                mapping[f"ATTRIBUTE_LABEL {i}"] = f"attributes[{i - 1}].label"
                mapping[f"ATTRIBUTE_VALUE {i}"] = f"attributes[{i - 1}].value"
                mapping[f"ATTRIBUTE_UOM {i}"] = f"attributes[{i - 1}].uom"

        elif alias:
            mapping[alias] = field_name

    return mapping


OUTPUT_HEADERS = _get_output_headers()
OUTPUT_HEADER_TO_FIELD = _get_output_header_to_field_map()


def _extract_value_from_field(field_data: Any) -> str:
    """Extract the actual value from an ExtractedField or ExtractedAttribute dict."""
    if field_data is None:
        return ""

    if isinstance(field_data, dict):
        if "value" in field_data:
            val = field_data["value"]

            if isinstance(val, dict) and "value" in val:
                return str(val["value"]) if val["value"] is not None else ""

            return str(val) if val is not None else ""

        if "label" in field_data and "value" in field_data:
            val = field_data["value"]

            if isinstance(val, dict) and "value" in val:
                return str(val["value"]) if val["value"] is not None else ""

            return str(val) if val is not None else ""

    return str(field_data)


def _get_nested_value(data: dict, path: str) -> Any:
    """Get value from nested dict using path like 'item_features[0]' or 'attributes[0].label'."""
    current = data

    parts = path.replace("]", "").split("[")

    for part in parts:
        if "." in part:
            key, subkey = part.split(".", 1)

            if key.isdigit():
                key = int(key)

            # Handle list access
            if isinstance(current, list) and isinstance(key, int) and 0 <= key < len(current):
                current = current[key]

            elif isinstance(current, dict):
                current = current.get(key, {})

            else:
                return None

            # Now get subkey
            if isinstance(current, dict):
                current = current.get(subkey, {})

            else:
                return None

        else:
            if part.isdigit():
                part = int(part)

            if isinstance(current, list) and isinstance(part, int) and 0 <= part < len(current):
                current = current[part]

            elif isinstance(current, dict):
                current = current.get(part, {})

            else:
                return None

    return current


def _build_output_row_data(
    row: JobRow,
    output_headers: list[str],
    header_to_field: dict[str, str],
) -> list[str]:
    """Build a single output row with values mapped from result_data to Output headers."""
    result_data = row.result_data or {}
    row_values: list[str] = []

    for header in output_headers:
        field_path = header_to_field.get(header)

        if not field_path:
            row_values.append("")
            continue

        if field_path.startswith("item_features[") or field_path.startswith("attributes["):
            value = _get_nested_value(result_data, field_path)

            if field_path.endswith(".value"):
                row_values.append(_extract_value_from_field(value))

            else:
                # For item_features[index] (not .value), extract the ExtractedField value
                row_values.append(_extract_value_from_field(value))

        else:
            field_data = result_data.get(field_path)
            row_values.append(_extract_value_from_field(field_data))

    return row_values


def _collect_conflicts(rows: list[JobRow]) -> list[dict]:
    """Collect all conflicts from result_data across all rows."""
    conflicts: list[dict] = []

    for row in rows:
        if not row.result_data:
            continue

        result_data = row.result_data

        for field_name, field_data in result_data.items():
            if isinstance(field_data, dict) and field_data.get("is_conflict"):
                conflict_details = field_data.get("conflict_details")

                if conflict_details:
                    conflict_row = {
                        "row_number": row.row_number,
                        "field": conflict_details.get("field", field_name),
                        "selected_value": conflict_details.get("selected_value", ""),
                        "selected_uom": conflict_details.get("selected_uom", ""),
                        "selected_source": conflict_details.get("selected_source", ""),
                        "conflicting_value": conflict_details.get("conflicting_value", ""),
                        "conflicting_uom": conflict_details.get("conflicting_uom", ""),
                        "conflicting_source": conflict_details.get("conflicting_source", ""),
                        "recommendation": conflict_details.get("recommendation", ""),
                    }

                    conflicts.append(conflict_row)

            if field_name == "attributes" and isinstance(field_data, list):
                for attr in field_data:
                    if isinstance(attr, dict):
                        attr_value = attr.get("value")

                        if isinstance(attr_value, dict) and attr_value.get("is_conflict"):
                            conflict_details = attr_value.get("conflict_details")

                            if conflict_details:
                                label = attr.get("label", "")

                                conflict_row = {
                                    "row_number": row.row_number,
                                    "field": f"{conflict_details.get('field', 'attribute')}: {label}",
                                    "selected_value": conflict_details.get("selected_value", ""),
                                    "selected_uom": conflict_details.get("selected_uom", ""),
                                    "selected_source": conflict_details.get("selected_source", ""),
                                    "conflicting_value": conflict_details.get("conflicting_value", ""),
                                    "conflicting_uom": conflict_details.get("conflicting_uom", ""),
                                    "conflicting_source": conflict_details.get("conflicting_source", ""),
                                    "recommendation": conflict_details.get("recommendation", ""),
                                }

                                conflicts.append(conflict_row)

    return conflicts


def _copy_cell_style(source_cell, target_cell) -> None:
    """Copy style from source cell to target cell."""
    if source_cell.has_style:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color,
        )

        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color,
        )

        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom,
        )

        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            wrap_text=source_cell.alignment.wrap_text,
        )

        target_cell.number_format = source_cell.number_format


def _build_workbook(
    job: Job,
    rows: list[JobRow],
    input_headers: list[str],
) -> openpyxl.Workbook:
    """Build the complete workbook with Input, Output, and Conflicts sheets."""
    original_path = Path(job.input_file_path)

    template_wb = None
    if job.input_format == "xlsx":
        if not original_path.exists():
            raise FileNotFoundError(
                f"Original input file not found: {original_path}"
            )
        template_wb = openpyxl.load_workbook(original_path)

    wb = openpyxl.Workbook()
    template_ws = (
        template_wb[template_wb.sheetnames[0]]
        if template_wb and template_wb.sheetnames
        else None
    )

    # Input sheet contains every current database row.
    ws_input = wb.active
    ws_input.title = "Input"
    input_sheet_headers = ["row_number"] + input_headers
    ws_input.append(input_sheet_headers)

    fallback_style_cell = None
    if template_ws is not None:
        for col_idx in range(1, template_ws.max_column + 1):
            source_cell = template_ws.cell(row=1, column=col_idx)
            if source_cell.has_style:
                fallback_style_cell = source_cell
                break

    for col_idx in range(1, len(input_sheet_headers) + 1):
        target_cell = ws_input.cell(row=1, column=col_idx)
        if template_ws is not None:
            source_cell = template_ws.cell(row=1, column=col_idx)
            if source_cell.has_style:
                _copy_cell_style(source_cell, target_cell)
                continue
        if fallback_style_cell is not None:
            _copy_cell_style(fallback_style_cell, target_cell)
            continue
        target_cell.font = Font(bold=True)

    for row in rows:
        ws_input.append(
            [
                row.row_number,
                *[
                    row.input_data.get(key, "") if row.input_data else ""
                    for key in input_headers
                ],
            ]
        )

    # Output contains only terminal rows with exportable result data.
    output_headers = ["row_number"] + OUTPUT_HEADERS
    ws_output = wb.create_sheet("Output")
    ws_output.append(output_headers)
    for cell in ws_output[1]:
        cell.font = Font(bold=True)

    for row in rows:
        if row.status not in ("completed", "failed") or not row.result_data:
            continue
        ws_output.append(
            [
                row.row_number,
                *_build_output_row_data(
                    row,
                    OUTPUT_HEADERS,
                    OUTPUT_HEADER_TO_FIELD,
                ),
            ]
        )

    # Conflicts is always present and reflects the current result data.
    conflict_headers = [
        "row_number",
        "field",
        "selected_value",
        "selected_uom",
        "selected_source",
        "conflicting_value",
        "conflicting_uom",
        "conflicting_source",
        "recommendation",
    ]
    ws_conflicts = wb.create_sheet("Conflicts")
    ws_conflicts.append(conflict_headers)
    for cell in ws_conflicts[1]:
        cell.font = Font(bold=True)
    for conflict in _collect_conflicts(rows):
        ws_conflicts.append([conflict.get(header, "") for header in conflict_headers])

    return wb


def generate_output(job_id: str) -> Path | None:
    """Generate final persistent output for completed jobs. Sets job.output_file_path."""
    with SessionLocal() as session:
        job = session.get(Job, job_id)

        if job is None:
            return None

        rows = session.execute(
            select(JobRow)
            .where(JobRow.job_id == job.id)
            .order_by(JobRow.row_number)
        ).scalars().all()

        if not rows:
            return None

        input_headers: list[str] = []
        seen: set[str] = set()

        for row in rows:
            if row.input_data:
                for key in row.input_data.keys():
                    if key not in seen:
                        seen.add(key)
                        input_headers.append(key)

        output_path = get_output_file_path(
            job_id,
            job.input_filename,
            job.input_format,
        )

        wb = _build_workbook(
            job,
            rows,
            input_headers,
        )

        wb.save(output_path)

        job.output_file_path = str(output_path)
        session.add(job)
        session.commit()

        return output_path


def generate_partial_output(job_id: str) -> Path | None:
    """Generate a partial output for download during job processing. Returns a temp file path."""
    with SessionLocal() as session:
        job = session.get(Job, job_id)

        if job is None:
            return None

        rows = session.execute(
            select(JobRow)
            .where(JobRow.job_id == job.id)
            .order_by(JobRow.row_number)
        ).scalars().all()

        if not rows:
            return None

        has_terminal = any(
            row.status in ("completed", "failed")
            for row in rows
        )

        if not has_terminal:
            return None

        input_headers: list[str] = []
        seen: set[str] = set()

        for row in rows:
            if row.input_data:
                for key in row.input_data.keys():
                    if key not in seen:
                        seen.add(key)
                        input_headers.append(key)

        suffix = ".xlsx"

        with tempfile.NamedTemporaryFile(
            suffix=suffix,
            delete=False,
        ) as tmp:
            temp_path = Path(tmp.name)

        wb = _build_workbook(
            job,
            rows,
            input_headers,
        )

        wb.save(temp_path)

        return temp_path


if __name__ == "__main__":
    pytest.main([__file__, "-v"])